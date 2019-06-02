
#coding=utf-8
import xlrd
from xlutils.copy import copy

Python总结——处理Excel（xlrd,openpyxl 模块）
2016-05-04  橙子 签约作者
Python总结——处理Excel（xlrd,openpyxl 模块）

(原创，如要转载请标明出处及作者) 
Python处理Excel常用操作就是读和写，我的需求是需要在原excel文件中进行读写操作。共用到了两个模块xlrd和openpyxl,这两个模块都是需要自己去安装的。openpyxl只能用来处理 Excel 2007 及以上版本的 excel 文件，也就是 .xlsx/.xlsm 格式的表格文件

xlrd模块

使用步骤及方法：

打开文件：

import xlrd

excel=xlrd.open_workbook('E:/test.xlsx')

 

获取sheet：

table = excel.sheets()[0]       #通过索引获取  

table = excel.sheet_by_index(0)    #通过索引获取  

table = excel.sheet_by_name('Sheet1')   #通过表名获取  

 

备注：以下方法的操作都要在sheet基础上使用

获取行数和列数：

rows=table.nrows   #获取行数

cols=table.ncols    #获取列数

 

获取单元格值：

Data=table.cell(row,col).value  #获取表格内容，是从第一行第一列是从0开始的，注意不要丢掉 .value

 

获取整行或整列内容

Row_values=table.row_values(i)   #获取整行内容

Col_values=table.col_values(i)   #获取整列内容

 

Openpyxl模块

 

读

打开文件：

from openpyxl import load_workbook

excel=load_workbook('E:/test.xlsx')

 

获取sheet：

table = excel.get_sheet_by_name('Sheet1')   #通过表名获取  

 

获取行数和列数：

rows=table.max_row   #获取行数

cols=table.max_column    #获取列数

 

获取单元格值：

Data=table.cell(row=row,column=col).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value

 

写

 

#打开文件

excel=load_workbook('E:/test.xlsx')

sheet=excel.active

#设定单元格的值，两种方式

sheet.cell(row=2,column=5).value=99

sheet.cell(row=3,column=5,value=100)

#保存修改的文件

excel.save('E:/test.xlsx')