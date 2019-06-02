#coding=utf-8
import shutil,xlrd,os
from xlutils.copy import copy
import openpyxl
#xlwd


"""
	只支持.xlsx文件
"""

class OperaExcel:
	def __init__(self,file_path=None,i=None):
		if file_path == None:
			self.file_path = '..\\config\\testcase.xlsx'
		else:
			self.file_path = file_path	
		if i == None:
			i=0
		self.excel = self.get_excel()
		self.data = self.get_sheet(i)
			
	def get_excel(self):
		'''
		打开并获取，excel对象
		'''
		#file_path_end="../02.config/testcase_temp.xlsx"
		#shutil.copyfile(self.file_path,file_path_end)
		excel = openpyxl.load_workbook(self.file_path)
		return excel

	def get_sheet(self,i):
		'''
		获取一个i键(或index)对应的sheet页对象
		参数：
			--i：sheet页名或者sheet的键值
		'''
		if isinstance(i,int):
			tables = self.excel.worksheets[i]
		elif isinstance(i,str):
			table = self.excel.get_sheet_by_name(i)
		return tables

	def get_rows(self):
		'''
		获取sheet页的行数
		'''
		rows = self.data.max_row
		return rows

	def get_columns(self):
		'''
		获取sheet页的行数
		'''
		cols = self.data.max_column
		return cols

	def get_cell(self,row,column):
		'''
		获取sheet页单元格的值
		参数：
			--row：所在行
			--column：所在列
		'''
		data = self.data.cell(row,column).value
		return data

	def write_value(self,row_name="",column_name="实际结果",value=""):
		'''
		向sheet页单元格写入
		参数：
			--row_name：行名，即一行的第一个单元格的值
				根据案例模板是Case_ID列的值，通过该值获取对应Case_ID所在的行数
			--column_name：列名，即一列的第一个单元格的值
				通过该值获取对应的列数,结合row_name获得的行数，修改对应单元格的值
				根据案例模板多为：实际结果、是否执行、是否通过
		Etc：
			修改案例 "test_04" 的 执行状态 为 "pass"
			OperaExcel().write_value("test_04","执行状态","通过")
		'''
		excel=OperaExcel().get_excel()
		sheet = excel.worksheets[0]
		for i in range(1,sheet.max_row+1):
			if sheet.cell(i,1).value==row_name:
				row=i
				break
		for i in range(1,sheet.max_column+1):
			if sheet.cell(1,i).value==column_name:
				column=i
				break
		sheet.cell(row,column).value=value
		excel.save(self.file_path)


if __name__ == '__main__':
	for i in range(1000):
		OperaExcel().write_value("test_04","实际结果","实际结果")
	



